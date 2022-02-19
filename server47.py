import time
from datetime import datetime
import traceback # stack trace
import os
from td.client import TDClient
import openpyxl as excel # excel library
from data import TD_ACCOUNT, CONSUMER_KEY, REDIRECT_URI, JSON_PATH # Import from data.py




#pip install td-ameritrade-python-api
#https://www.youtube.com/watch?v=8N1IxYXs4e8

#=====================================================================================================================================================================
#           TD AMERITRADE
#=====================================================================================================================================================================
TD_CLIENT = TDClient(client_id= CONSUMER_KEY, redirect_uri= REDIRECT_URI, credentials_path= JSON_PATH)
TD_CLIENT.login() # Authenticate. This needs to be done every 90 days.

# get the transactions that are BUY ORDERS ONLY
TRANSACTIONS_DICT = TD_CLIENT.get_transactions(account=TD_ACCOUNT, transaction_type='BUY_ONLY') 

# get the transactions that were dividends
DIVIDENDS_DICT = TD_CLIENT.get_transactions(account=TD_ACCOUNT, transaction_type='DIVIDEND') 
PRINCIPAL = TD_CLIENT.get_transactions(account=TD_ACCOUNT, transaction_type='CASH_IN_OR_CASH_OUT')
# get the account data
ACCOUNT_DATA_DICT = TD_CLIENT.get_accounts(account=TD_ACCOUNT, fields=['orders'])

# position data
POSITION_DATA_DICT = TD_CLIENT.get_accounts(account=TD_ACCOUNT, fields=["positions"])

# account value of TD Ameritrade account
ACCOUNT_VALUE = ACCOUNT_DATA_DICT['securitiesAccount']['initialBalances']['accountValue']

def get_owned_position_symbols() -> list:
    '''
    Get a list of all the SYMBOLs of the positions that are held
    '''
    owned_positions = []
    for i in range(0, len(POSITION_DATA_DICT["securitiesAccount"]["positions"])):
        if (str(POSITION_DATA_DICT["securitiesAccount"]["positions"][i]["instrument"]["symbol"]) != "MMDA1" ): # exclude MMDA1 (money stored on account, but not invested)
            owned_positions.append(POSITION_DATA_DICT["securitiesAccount"]["positions"][i]['instrument']['symbol'])
            #print(str(POSITION_DATA_DICT["securitiesAccount"]["positions"][i]['instrument']['symbol']))
    return owned_positions

#=====================================================================================================================================================================
#           EXCEL WORK BOOK
#=====================================================================================================================================================================
# | FIELDS | #
BASE_WORK_BOOK_PATH = "./excelworkbook/base.xlsx"
SAVE_PATH = "./portfolio.xlsx"

# if portfolio.xlsx exists, use it as the main file, otherwise use the base.xlsx file
if os.path.exists("./portfolio.xlsx"):
    BASE_WORK_BOOK_PATH = "./portfolio.xlsx"
    EXCEL_WORK_BOOK = excel.load_workbook("./portfolio.xlsx")
else:
    EXCEL_WORK_BOOK = excel.load_workbook(BASE_WORK_BOOK_PATH)

ws_transactions = EXCEL_WORK_BOOK['Transactions']     # Transactions worksheet in TD Ameritrade Stonks.xlxs
ws_contributed = EXCEL_WORK_BOOK['Contributed']     # $Contributed$ worksheet in TD Ameritrade Stonks.xlxs
ws_portfolio = EXCEL_WORK_BOOK['Portfolio']           # Portfolio worksheet in TD Ameritrade Stonks.xlxs
ws_position_data = EXCEL_WORK_BOOK['Position Data']   # Position Data worksheet in TD Ameritrade Stonks.xlxs
ws_dividends = EXCEL_WORK_BOOK['Dividends']


'''
Update the 'Position Data' excel worksheet with all owned symbol data (bid price, last price, etc...)
'''
def update_stock_data() -> None:
    quotes = TD_CLIENT.get_quotes(get_owned_position_symbols())
    values = list(quotes.values())      # list of all dictionary key values (bidPrice, etc.)
    index = 1
    for stonk in values:
        index+=1
        
        if (ws_position_data.cell(row=index, column=4).value != None):
            old_price = float(ws_position_data.cell(row=index, column=4).value)
            new_price = float(stonk['lastPrice'])
            delta = ((new_price - old_price) / old_price)
            percent = "{:.2%}".format(delta) # format 2 decimal places example: 0.0345 = 3.45
            print(f"{stonk['symbol']} Last Price changed from: ${(ws_position_data.cell(row=index, column=4).value)} to: ${stonk['lastPrice']} | change: {percent}")

        ws_position_data.cell(row=index, column=1, value=str(stonk['symbol'])).number_format = '$#,##0.00'          # update symbol cells
        ws_position_data.cell(row=index, column=2, value=float(stonk['bidPrice'])).number_format = '$#,##0.00'      # update bidPrice cells
        ws_position_data.cell(row=index, column=3, value=float(stonk['askPrice'])).number_format = '$#,##0.00'      # update askPrice cells
        ws_position_data.cell(row=index, column=4, value=float(stonk['lastPrice'])).number_format = '$#,##0.00'     # update lastPrice cells
        ws_position_data.cell(row=index, column=5, value=float(stonk['openPrice'])).number_format = '$#,##0.00'     # update openPrice cells
        ws_position_data.cell(row=index, column=6, value=float(stonk['highPrice'])).number_format = '$#,##0.00'     # update highPrice cells
        ws_position_data.cell(row=index, column=7, value=float(stonk['lowPrice'])).number_format = '$#,##0.00'      # update lowPrice cells
        ws_position_data.cell(row=index, column=8, value=float(stonk['closePrice'])).number_format = '$#,##0.00'    # update closePrice cells

    EXCEL_WORK_BOOK.save(SAVE_PATH)


'''
Update the 'Dividends' excel worksheet with all dividend transactions
'''
def update_dividend_data() -> None:
    try:
        print(f"Updated the {ws_dividends} with the following transaction data: ")
        dividend_list = list(DIVIDENDS_DICT) 
        dividend_list.reverse()
        
        for i in range(0, len(dividend_list)):
            transaction_id = int(dividend_list[i]['transactionId']) # grab the ID from the list
            symbol = str(dividend_list[i]['transactionItem']['instrument']['symbol'])
            date = convert_annoying_date_format(str((dividend_list[i]['transactionDate']))) # convert the annoying tedious ameritradious date format to a readable format.
            amount_recieved = float(dividend_list[i]['netAmount'])
            

            
            rows = list(ws_dividends['A']) # list of rows in the 'A' column
            for index in range(2, len(rows)+2): # start iterating at the row below the column header
                # if the entry is already found, don't overwrite the cell and break the loop
                if (ws_dividends.cell(row=index, column=1).value == transaction_id):
                    break
                # if the row already has data populated
                if (ws_dividends.cell(row=index, column=1).value != None):
                    continue # iterate the loop
                ws_dividends.cell(row=index, column=1, value=transaction_id)         # update ID cells
                ws_dividends.cell(row=index, column=2, value=symbol)     # update symbol cells
                ws_dividends.cell(row=index, column=3, value=date)       # update date cells
                ws_dividends.cell(row=index, column=4, value=amount_recieved).number_format = '$#,##0.00' # amount recieved for the dividend
                print(f"ID: {transaction_id} SYMBOL: {symbol} DATE: {date} DIV AMOUNT: ${amount_recieved}")
                break

        # save the excel file
        EXCEL_WORK_BOOK.save(SAVE_PATH)
    except Exception as e:
        traceback.print_exc() # stacktrace
        input("There was an error updating the transactions. Make sure the spreadsheet isn't already opened. Press Enter to continue...")

'''
Update the 'Contributed' excel worksheet with all principal transactions
'''
def update_contributed_data() -> None:
    try:
        print(f"Updated the {ws_contributed} with the following transaction data: ")
        
        # Append the dictionary to a list and reverse it so the transactions show in ascending order
        principal_list = list(PRINCIPAL) 
        principal_list.reverse()

        # for i = 0; i < the size of the list; i++
        for i in range(0, len(principal_list)):               
            transaction_id = int(principal_list[i]['transactionId']) # grab the ID from the list
            date = convert_annoying_date_format(str((principal_list[i]['transactionDate']))) # convert the annoying tedious ameritradious date format to a readable format.
            amount_contributed = float(principal_list[i]['netAmount'])
            
            rows = list(ws_contributed['A']) # list of rows in the 'A' column that have data
            for index in range(2, len(rows)+2):
                # if the entry is already found, don't overwrite the cell
                if (ws_contributed.cell(row=index, column=1).value == transaction_id):
                    break
                # if the row already has data populated
                if (ws_contributed.cell(row=index, column=1).value != None):
                    continue # iterate the loop
                ws_contributed.cell(row=index, column=1, value=transaction_id)         # update ID cells
                ws_contributed.cell(row=index, column=2, value=date)                   # update date cells
                ws_contributed.cell(row=index, column=3, value=amount_contributed).number_format = '$#,##0.00' # amount contributed to the TD account
                print(f"ID: {transaction_id} DATE: {date} AMOUNT: ${amount_contributed}")
                break
            
        # save the excel file
        EXCEL_WORK_BOOK.save(SAVE_PATH)
    except Exception as e:
        traceback.print_exc() # stacktrace
        input("There was an error updating the transactions. Make sure the spreadsheet isn't already opened. Press Enter to continue...")


def update_portfolio() -> None:
    # Iterate through the columns and rows to look for the cell that has "Symbol" as the value
    for column in range(1, 50):
        for row in range(1, 50):
            if (ws_portfolio.cell(row=row, column=column).value == "Symbol"): # If the cell has 'Symbol' as the value
                for symbol in get_owned_position_symbols(): # For every symbol that is owned
                    row=row+1 # iterate to the row just below the "Symbol" column
                    ws_portfolio.cell(row=row, column=column, value=symbol) # Update the cells below the "Symbol" cell
                break # Break out of the loop as we've finished updating
    print(f"Updated {ws_portfolio} 'Symbols' to: {get_owned_position_symbols()}")
    EXCEL_WORK_BOOK.save(SAVE_PATH)
                

def update_account_value() -> None:
    """
    Update total Account Value
    """
    currentYear = datetime.now()
    year = int(currentYear.date().strftime("%Y"))
    try:

        # Iterate through the columns and rows
        for column in range(1, 50):
            for row in range(1, 50):
                if (ws_portfolio.cell(row=row, column=column).value == "Account Value"): # If the cell has 'Account Value' as the value
                    print(f"Updated {ws_portfolio} 'Account Value' from: {ws_portfolio.cell(row=row+1, column=column).value} to: ${ACCOUNT_VALUE}")
                    ws_portfolio.cell(row=row+1, column=column, value=ACCOUNT_VALUE) # Update the cell just below "Account Value" cell
                
        # if (year == 2021):
        #     print(f"Updated {ws_contributed} 'Account Value' from: {ws_contributed.cell(row=2, column=4).value} to: ${ACCOUNT_VALUE}")
        #     ws_contributed.cell(row=2, column=4, value=ACCOUNT_VALUE)    # update YTD account value cell for 2021
        # if (year == 2022):
        #     print(f"Updated {ws_contributed} 'Account Value' from: {ws_contributed.cell(row=2, column=9).value} to: ${ACCOUNT_VALUE}")
        #     ws_contributed.cell(row=2, column=9, value=ACCOUNT_VALUE)    # update YTD account value cell for 2022

        # save the excel file
        EXCEL_WORK_BOOK.save(SAVE_PATH)
    except:
        #print(e)
        traceback.print_exc() # stacktrace
        input("There was an error updating the total account value. Make sure the spreadsheet isn't already opened. Press Enter to continue...")





def convert_annoying_date_format(date: str) -> str:
    '''
    Convert the date format of example: 2021-08-13T16:20:10+0000 -> 08/13/2021
    '''
    oldDateFormat = datetime.strptime(date, "%Y-%m-%dT%H:%M:%S%z")
    parsedDate = datetime.strftime(oldDateFormat, "%m/%d/%Y")
    return parsedDate
        

def update_transactions():
    '''
    Update the 'Transactions' sheet with data from every transaction in the 'transactions' dictionary
    '''
    try:
        print(f"Updated the {ws_transactions} with the following transaction data: ")
        transaction_list = list(TRANSACTIONS_DICT) 
        transaction_list.reverse()
        # for i = 0; i < the size of the transactions list; ++i
        for i in range(0, len(transaction_list)):    
            id = int(transaction_list[i]['orderId'])            # grab the ID from the list
            symbol = str(transaction_list[i]['transactionItem']['instrument']['symbol'])
            date = convert_annoying_date_format(str((transaction_list[i]['transactionDate']))) # convert the annoying tedious ameritradious date format to a readable format.
            amount_paid = float(transaction_list[i]['netAmount'])
            share_price = float(transaction_list[i]['transactionItem']['price'])
            shares = int(transaction_list[i]['transactionItem']['amount'])

            if (amount_paid < 0): # If the transaction was a "Buy"
                amount_paid = -1 * amount_paid # Convert the negative 'transaction' to a positive float.
            else:
                shares = -1 * shares # subtract the amount of 'shares' you own



            rows = list(ws_transactions['A']) # list of cells in the 'A' column that have data in them
            for index in range(2, len(rows)+2):
                # if the entry is already found, don't overwrite the cell
                if (ws_transactions.cell(row=index+1, column=1).value == id):
                    break
                # if the row already has data populated
                if (ws_transactions.cell(row=index, column=1).value != None):
                    continue # iterate the loop
                ws_transactions.cell(row=index, column=1, value=id)         # update ID cells
                ws_transactions.cell(row=index, column=2, value=symbol)     # update symbol cells
                ws_transactions.cell(row=index, column=3, value=date)       # update date cells
                ws_transactions.cell(row=index, column=4, value=share_price).number_format = '$#,##0.00' # update share_price cells
                ws_transactions.cell(row=index, column=5, value=shares)     # update shares cells
                ws_transactions.cell(row=index, column=6, value=amount_paid).number_format = '$#,##0.00' # update amount cells
                print(f"ID: {id} SYMBOL: {symbol} DATE: {date} SHARE PRICE: {share_price} SHARES: {shares} AMOUNT PAID: {amount_paid}")
                break
            
        # save the excel file
        EXCEL_WORK_BOOK.save(SAVE_PATH)
    except:
        traceback.print_exc() # stacktrace
        input("There was an error updating the transactions. Make sure the spreadsheet isn't already opened. Press Enter to continue...")




def recursive_update():
    update_stock_data()
    print("Updated Transactions! \n")
    time.sleep(10)
    recursive_update() #recursion



def main():
    currentYear = datetime.now()
    year = str(currentYear.date().strftime("%Y"))
    print("Do you want to run continuously? (Press 1), otherwise, press (Enter)...")
    non_stop_run = input().lower()
    if non_stop_run == "1":
        recursive_update()
    else:
        update_transactions()
        update_dividend_data()
        update_account_value()
        update_stock_data()
        update_contributed_data()
        
        print("Do you want to update the symbols you own? (y/n), otherwise, press (Enter)...")
        update_portfolio_symbols = input().lower()
        if update_portfolio_symbols == "y":
            update_portfolio()
        
        #EXCEL_WORK_BOOK.save(f"./excelworkbooktest/{year}.xlsx")
        input("Updated! Press Enter to close...")

if __name__ == "__main__":
    main()


